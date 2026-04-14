import openpyxl
import json
import re

file_1 = r"C:\Users\iamra\Downloads\final.xlsx"
file_2 = r"C:\Users\iamra\Downloads\Placement_Patent_Competetion Details (1).xlsx"

def clean_count(val):
    if val is None: return 0
    s_val = str(val).strip().lower()
    if s_val in ['no', '-', 'nil', 'none', 'nii', '0']: 
        return 0
    if s_val == 'yes':
        return 1
    
    if ',' in s_val:
        items = [i.strip() for i in s_val.split(',') if i.strip()]
        return len(items)
    
    match = re.search(r'\d+', s_val)
    if match:
        return int(match.group())
    
    if len(s_val) > 1:
        return 1
    return 0

def extract_data():
    # 1. Load Backlog and Achievement details from File 2 (Cleaner numerical data)
    wb2 = openpyxl.load_workbook(file_2, data_only=True)
    sheet2 = wb2.active
    backlog_data = {}
    
    for row in sheet2.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue # Skip if no Reg No
        reg_id = str(row[1]).strip()
        backlog_data[reg_id] = {
            "backlogs": 1 if str(row[5]).lower() == 'yes' else 0,
            "certs": clean_count(row[6]),
            "winning_hacks": clean_count(row[7]),
            "patents": clean_count(row[8])
        }

    # 2. Load Core Student Data (CGPA, Link, Skills) from File 1
    wb1 = openpyxl.load_workbook(file_1, data_only=True)
    sheet1 = wb1.active
    
    students = []
    # Skip header row
    for row in sheet1.iter_rows(min_row=2, values_only=True):
        if not row[2]: continue
        
        reg_no = str(row[2]).strip()
        name = str(row[1])
        
        # CGPA
        cgpa = 0.0
        try:
            cgpa_val = row[5] or row[4] or 0
            cgpa = round(float(cgpa_val), 2)
        except: pass
        
        # Skills
        domain = str(row[10] or "")
        skills = []
        if domain and domain.lower() != 'nil':
            skills = [s.strip() for s in re.split(r'[,/&]', domain) if s.strip()]
        
        if not skills:
            skills = ["Python", "Java", "DSA"]
            
        # Merge with Backlog Data
        details = backlog_data.get(reg_no, {"backlogs": 0, "certs": 0, "winning_hacks": 0, "patents": 0})
        
        # Prefer numeric counts from File 2 if available
        certs_count = details['certs'] if details['certs'] > 0 else clean_count(row[11])
        hack_count = details['winning_hacks'] if details['winning_hacks'] > 0 else clean_count(row[7])

        student = {
            "id": reg_no,
            "name": name,
            "dept": "CSE",
            "cgpa": cgpa,
            "total_backlogs": details['backlogs'], 
            "active_backlogs": details['backlogs'],
            "skills": skills,
            "hackathons": hack_count,
            "certifications": certs_count,
            "patents": details['patents'],
            "resume_link": str(row[12] or ""),
            "githubLink": str(row[9] or ""),
            "domain": domain
        }
        students.append(student)
    
    return students

if __name__ == "__main__":
    students = extract_data()
    print(json.dumps(students))
