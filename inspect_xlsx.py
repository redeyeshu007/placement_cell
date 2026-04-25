import openpyxl
import json

file_path = r"D:\Hackverse\hACKVERSE.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("--- First 5 rows from 'Students' sheet ---")
    if 'Students' in wb.sheetnames:
        sheet = wb['Students']
        for idx, row in enumerate(sheet.iter_rows(max_row=5, values_only=True)):
            print(f"Row {idx+1}: {row}")
    else:
        print("No 'Students' sheet found.")
except Exception as e:
    print(f"Error: {e}")
