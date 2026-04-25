import openpyxl
from pymongo import MongoClient

# Configuration
EXCEL_FILE = r"D:\Hackverse\hACKVERSE.xlsx"
MONGO_URI = "mongodb://localhost:27017/"
DB_NAME = "hackverse"

def migrate():
    try:
        # 1. Connect to MongoDB
        client = MongoClient(MONGO_URI)
        db = client[DB_NAME]
        students_col = db["students"]
        companies_col = db["companies"]
        
        # Clear existing data for a fresh start
        students_col.delete_many({})
        companies_col.delete_many({})
        print("Cleared existing collections.")

        # 2. Load Excel
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # --- MIGRATE STUDENTS ---
        if 'Students' in wb.sheetnames:
            sheet = wb['Students']
            headers = [cell.value for cell in sheet[1]]
            
            student_docs = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue # skip empty
                
                # Zip headers with row values to create a dict
                doc = dict(zip(headers, row))
                
                # Cleanup data for the Mongoose schema
                # Match the keys to the Mongoose model (src/models/Student.ts)
                cleaned_doc = {
                    "id": str(doc.get("id")),
                    "name": doc.get("name"),
                    "dept": doc.get("dept"),
                    "cgpa": float(doc.get("cgpa") or 0),
                    "totalBacklogs": int(doc.get("total_backlogs") or 0),
                    "activeBacklogs": int(doc.get("active_backlogs") or 0),
                    "skills": [s.strip() for s in str(doc.get("skills") or "").split(",") if s.strip()],
                    "hackathons": int(doc.get("hackathons") or 0),
                    "certifications": int(doc.get("certifications") or 0),
                    "resume_link": doc.get("resume_link"),
                    "githubLink": doc.get("githubLink"),
                    "domain": doc.get("domain")
                }
                student_docs.append(cleaned_doc)
            
            if student_docs:
                students_col.insert_many(student_docs)
                print(f"Successfully migrated {len(student_docs)} students.")
        
        # --- MIGRATE MOCK COMPANIES ---
        # Since companies were mock in frontend, I'll seed them here
        mock_companies = [
            { "id": "C001", "name": "Google India", "requiredSkills": ["Python", "Java", "DSA"], "prioritySkills": ["DSA", "System Design"], "minCgpa": 8.5, "allowedBacklogs": 0 },
            { "id": "C002", "name": "Razorpay", "requiredSkills": ["Node", "SQL", "Go"], "prioritySkills": ["Node"], "minCgpa": 8.0, "allowedBacklogs": 1 },
            { "id": "C003", "name": "Microsoft", "requiredSkills": ["C#", "Azure", "SQL"], "prioritySkills": ["Cloud Architecture"], "minCgpa": 8.5, "allowedBacklogs": 0 },
            { "id": "C004", "name": "TCS", "requiredSkills": ["Java", "C++", "SQL"], "prioritySkills": ["Java"], "minCgpa": 6.0, "allowedBacklogs": 2 },
            { "id": "C005", "name": "Zoho", "requiredSkills": ["C", "Web Development"], "prioritySkills": ["Problem Solving"], "minCgpa": 7.0, "allowedBacklogs": 0 }
        ]
        companies_col.insert_many(mock_companies)
        print("Successfully seeded mock companies.")

        client.close()
    except Exception as e:
        print(f"Error during migration: {e}")

if __name__ == "__main__":
    migrate()
